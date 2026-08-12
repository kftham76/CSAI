from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .models import Person, ValidationIssue
from .text_utils import clean_text, normalize_identity


DEFAULT_CLIENTS_ROOT = Path(r"D:\CSAI_CLIENTS")
EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}


def _issue(code: str, message: str, source: str = "rotation workbook") -> ValidationIssue:
    return ValidationIssue(code=code, message=message, source=source)


def discover_rotation_workbook(
    company_folder: str,
    circulation_year: int,
    clients_root: Path = DEFAULT_CLIENTS_ROOT,
) -> tuple[Path | None, list[ValidationIssue]]:
    agm_root = Path(clients_root) / company_folder / "AGM"
    search_dirs = (agm_root / str(circulation_year), agm_root)
    for directory in search_dirs:
        if not directory.is_dir():
            continue
        candidates = sorted(
            path
            for path in directory.iterdir()
            if path.is_file()
            and path.suffix.lower() in EXCEL_EXTENSIONS
            and ("retir" in path.name.lower() or "rotation" in path.name.lower())
        )
        if len(candidates) == 1:
            return candidates[0], []
        if len(candidates) > 1:
            return None, [
                _issue(
                    "rotation_workbook_ambiguous",
                    f"Multiple rotation workbooks were found in {directory}: "
                    + ", ".join(path.name for path in candidates),
                    str(directory),
                )
            ]
    return None, [
        _issue(
            "rotation_workbook_missing",
            f"No retirement/rotation workbook was found under {agm_root}.",
            str(agm_root),
        )
    ]


def _year_value(value) -> int | None:
    if isinstance(value, (date, datetime)):
        return value.year
    text = clean_text(value)
    if text.isdigit() and len(text) == 4:
        return int(text)
    return None


def _is_marked(value) -> bool:
    if value is None or value is False:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    text = clean_text(value).upper()
    return text not in {"", "N", "NO", "FALSE", "0", "-"}


def _sheet_matrix(path: Path) -> list[list[object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            if any(any(clean_text(value) for value in row) for row in rows):
                return rows
    finally:
        workbook.close()
    return []


def read_retiring_directors(
    path: Path,
    financial_year: int,
    directors: Iterable[Person],
) -> tuple[list[Person], list[ValidationIssue]]:
    try:
        rows = _sheet_matrix(path)
    except Exception as error:
        return [], [_issue("rotation_workbook_invalid", str(error), str(path))]
    year_positions: list[tuple[int, int]] = []
    for row_index, row in enumerate(rows):
        for column_index, value in enumerate(row):
            if _year_value(value) == financial_year:
                year_positions.append((row_index, column_index))
    if not year_positions:
        return [], [
            _issue(
                "rotation_year_missing",
                f"No {financial_year} column was found in the rotation workbook.",
                str(path),
            )
        ]
    if len(year_positions) > 1:
        return [], [
            _issue(
                "rotation_year_ambiguous",
                f"More than one {financial_year} column was found in the rotation workbook.",
                str(path),
            )
        ]
    header_row, year_column = year_positions[0]
    director_map = {normalize_identity(person.name): person for person in directors}
    selected: list[Person] = []
    unmatched: list[str] = []
    for row in rows[header_row + 1 :]:
        if year_column >= len(row) or not _is_marked(row[year_column]):
            continue
        possible_names = [clean_text(value) for value in row[:year_column] if clean_text(value)]
        match = None
        chosen_name = ""
        for candidate in possible_names:
            match = director_map.get(normalize_identity(candidate))
            if match:
                chosen_name = candidate
                break
        if match is None:
            chosen_name = possible_names[-1] if possible_names else "(blank name)"
            unmatched.append(chosen_name)
        elif all(normalize_identity(item.name) != normalize_identity(match.name) for item in selected):
            selected.append(match)
    if unmatched:
        return [], [
            _issue(
                "rotation_director_mismatch",
                "Checked rotation name(s) do not match current directors: " + ", ".join(unmatched),
                str(path),
            )
        ]
    if not selected:
        return [], [
            _issue(
                "rotation_unmarked",
                f"No retiring director is checked under the {financial_year} column.",
                str(path),
            )
        ]
    return selected, []
