"""Convert EBOS data Excel to JSON.
Uses openpyxl to preserve leading-zero strings (IC, Contact No).
"""

import json
import openpyxl
from pathlib import Path

EXCEL_PATH = Path(r"D:\CSAI_DATA\Database\Ebos data.xlsx")
JSON_PATH = Path(r"D:\CSAI_DATA\Database\Ebos data.json")


def _is_numeric(s):
    """Check if string represents a number."""
    try:
        float(s)
        return True
    except (ValueError, TypeError):
        return False


def _to_num(s):
    """Convert numeric string to int or float."""
    f = float(s)
    return int(f) if f == int(f) else f


def excel_to_json(excel_path, json_path):
    """Read each sheet directly via openpyxl, convert to JSON dict."""

    wb = openpyxl.load_workbook(excel_path)
    output = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        # First row is header
        header = [str(h) if h is not None else f"Col_{i}" for i, h in enumerate(rows[0])]

        records = []
        for row in rows[1:]:
            # Skip fully empty separator rows
            if all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row):
                continue

            rec = {}
            has_data = False
            for i, val in enumerate(row):
                key = header[i] if i < len(header) else f"Col_{i}"
                if val is None:
                    rec[key] = ""
                elif isinstance(val, float):
                    # Clean up float display: 40.0 -> 40, 40.5 -> 40.5
                    if val == int(val):
                        rec[key] = int(val)
                    else:
                        rec[key] = round(val, 6)
                elif isinstance(val, int):
                    rec[key] = val
                else:
                    val_str = str(val)
                    # Convert Criteria values to numbers
                    if "Criteria" in key and _is_numeric(val_str):
                        rec[key] = _to_num(val_str)
                    else:
                        rec[key] = val_str
                if rec[key] != "":
                    has_data = True

            if has_data:
                records.append(rec)

        output[sheet_name] = records

    wb.close()

    # Write JSON
    json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    total = sum(len(v) for v in output.values())
    print(f"Converted {len(output)} sheet(s), {total} record(s).")
    print(f"Output: {json_path}")


if __name__ == "__main__":
    excel_to_json(EXCEL_PATH, JSON_PATH)
