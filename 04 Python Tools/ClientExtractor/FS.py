"""Extract the latest financial-statement filing for every CSAI client.

The extractor prefers the latest financial year found inside each PDF.  When
several documents cover that year, approved/final filings take precedence over
lodged copies, previews, ordinary reports, and drafts.  The resulting workbook
is deliberately small: one auditable row per company with an FS folder.
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import date, datetime
import os
from pathlib import Path
import re
import unicodedata
import warnings
from typing import Iterable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from pypdf import PdfReader


CLIENT_ROOT = Path(r"D:\CSAI_CLIENTS")
OUTPUT_FILE = Path(r"D:\CSAI_DATA\Database\FS.Xlsx")
SHEET_NAME = "FS"

COLUMNS = [
    "Company",
    "Source PDF",
    "Company's current financial year start date",
    "Company's current financial year end date",
    "Date of financial statements approved by Board of Directors",
    "Date of circulation of financial statements and reports to members",
    "Date of Statutory Declaration",
    "Statutory Declaration - Name of director who made declaration",
    "Number of directors signing Statement by Directors",
    "Name of first director who signed Statement by Directors",
    "Name of second director who signed Statement by Directors",
    "Name of audit firm",
    "Director's remuneration - Fees (Current Financial Year)",
]

DATE_COLUMNS = COLUMNS[2:7]
COUNT_COLUMN = "Number of directors signing Statement by Directors"
FEE_COLUMN = "Director's remuneration - Fees (Current Financial Year)"

DATE_TOKEN = r"(?:\d{4}-\d{2}-\d{2}|\d{1,2}/\d{1,2}/\d{4}|\d{1,2}\s+[A-Za-z]+\s+\d{4})"

warnings.filterwarnings("ignore", category=UserWarning, module="pypdf")


@dataclass(frozen=True)
class Candidate:
    path: Path
    financial_year_end: date
    explicit_filing_date: bool
    completeness: int
    page_count: int


def normalize_text(value: str) -> str:
    """Normalize PDF text without destroying line or column boundaries."""
    value = unicodedata.normalize("NFKC", value or "")
    replacements = {
        "\ufb01": "fi",
        "\ufb02": "fl",
        "\u00a0": " ",
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
        "\u2013": "-",
        "\u2014": "-",
    }
    for source, target in replacements.items():
        value = value.replace(source, target)
    return "\n".join(line.rstrip() for line in value.splitlines())


def flat_text(value: str) -> str:
    return re.sub(r"\s+", " ", normalize_text(value)).strip()


def parse_date_token(value: str | None) -> date | None:
    if not value:
        return None
    value = re.sub(r"\s+", " ", value.strip())
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d %B %Y", "%d %b %Y"):
        try:
            return datetime.strptime(value, pattern).date()
        except ValueError:
            continue
    return None


def date_after_label(text: str, label_pattern: str) -> date | None:
    match = re.search(
        rf"{label_pattern}\s*[:\-]?\s*({DATE_TOKEN})",
        flat_text(text),
        re.IGNORECASE,
    )
    return parse_date_token(match.group(1)) if match else None


def extract_financial_year_end(text: str) -> tuple[date | None, bool]:
    explicit = date_after_label(
        text,
        r"Company's\s+current\s+financial\s+year\s+end\s+date",
    )
    if explicit:
        return explicit, True

    conventional = re.search(
        rf"(?:financial\s+year|year)\s+ended\s+({DATE_TOKEN})",
        flat_text(text),
        re.IGNORECASE,
    )
    if conventional:
        return parse_date_token(conventional.group(1)), False
    return None, False


def _ocr_engine():
    """Create the explicitly requested OCR engine."""
    try:
        import numpy as np  # type: ignore
        import pypdfium2  # type: ignore
        from rapidocr_onnxruntime import RapidOCR  # type: ignore

        return pypdfium2, np, RapidOCR()
    except ImportError as error:
        raise RuntimeError(
            "OCR was requested but its optional dependencies are unavailable. "
            "Install numpy, pypdfium2, rapidocr-onnxruntime and onnxruntime, "
            "or rerun without --ocr."
        ) from error


_OCR = None
_OCR_CHECKED = False


def get_ocr_engine():
    global _OCR, _OCR_CHECKED
    if not _OCR_CHECKED:
        _OCR = _ocr_engine()
        _OCR_CHECKED = True
    return _OCR


def ocr_pages(pdf_path: Path, page_numbers: Iterable[int]) -> dict[int, str]:
    page_numbers = list(page_numbers)
    if not page_numbers:
        return {}

    pypdfium2, np, engine = get_ocr_engine()
    output: dict[int, str] = {}
    document = pypdfium2.PdfDocument(str(pdf_path))
    for position, page_number in enumerate(page_numbers, start=1):
        if page_number < 0 or page_number >= len(document):
            continue
        print(
            f"    OCR: {pdf_path.name} page {page_number + 1}/{len(document)} "
            f"({position}/{len(page_numbers)} sparse page(s))"
        )
        try:
            image = document[page_number].render(scale=2.0).to_pil()
            result, _ = engine(np.asarray(image))
            lines = [item[1] for item in (result or []) if len(item) > 1]
            output[page_number] = normalize_text("\n".join(lines))
        except Exception:
            continue
    return output


def read_head(pdf_path: Path, allow_ocr: bool = False) -> tuple[str, int]:
    reader = PdfReader(str(pdf_path))
    pages = [normalize_text(page.extract_text() or "") for page in reader.pages[:3]]
    combined = "\n".join(pages)
    if allow_ocr and len(flat_text(combined)) < 150:
        recovered = ocr_pages(pdf_path, range(min(3, len(reader.pages))))
        for index, text in recovered.items():
            if len(flat_text(text)) > len(flat_text(pages[index])):
                pages[index] = text
        combined = "\n".join(pages)
    return combined, len(reader.pages)


def candidate_stage(candidate: Candidate) -> tuple:
    name = candidate.path.name.casefold()
    if "receipt" in name or re.search(r"\bor[_ -]?xb", name):
        stage = 9
    elif "draft" in name:
        stage = 5
    elif "approved" in name or "final" in name:
        stage = 0
    elif "view filing" in name or name.startswith("fs-") or "lodge" in name:
        stage = 1
    elif "preview" in name:
        stage = 2
    elif "independent auditor" in name and not candidate.explicit_filing_date:
        stage = 8
    else:
        stage = 3

    try:
        modified = -candidate.path.stat().st_mtime
    except OSError:
        modified = 0
    return (
        stage,
        0 if candidate.explicit_filing_date else 1,
        -candidate.completeness,
        -candidate.page_count,
        modified,
        str(candidate.path).casefold(),
    )


def make_candidate(pdf_path: Path, allow_ocr: bool = False) -> Candidate | None:
    try:
        head, page_count = read_head(pdf_path, allow_ocr=allow_ocr)
    except Exception as error:
        print(f"  WARNING: cannot read {pdf_path.name}: {error}")
        return None

    financial_year_end, explicit = extract_financial_year_end(head)
    if not financial_year_end:
        return None

    searchable = flat_text(head)
    markers = (
        "current financial year start date",
        "approved by board of directors",
        "date of statutory declaration",
        "financial statements and reports",
    )
    completeness = sum(marker in searchable.casefold() for marker in markers)
    return Candidate(
        path=pdf_path,
        financial_year_end=financial_year_end,
        explicit_filing_date=explicit,
        completeness=completeness,
        page_count=page_count,
    )


def find_fs_folder(company_folder: Path) -> Path | None:
    try:
        return next(
            child
            for child in company_folder.iterdir()
            if child.is_dir() and child.name.casefold() == "fs"
        )
    except (StopIteration, OSError):
        return None


def is_receipt_pdf(pdf_path: Path) -> bool:
    name = pdf_path.name.casefold()
    return "receipt" in name or bool(
        re.search(r"(?:^|[^a-z0-9])or[_ -]?xb", name)
    )


def select_latest_pdf(
    fs_folder: Path,
    allow_ocr: bool = False,
) -> Candidate | None:
    pdfs = sorted(
        (
            path
            for path in fs_folder.rglob("*")
            if path.is_file() and path.suffix.casefold() == ".pdf"
            and not is_receipt_pdf(path)
        ),
        key=lambda path: str(path).casefold(),
    )
    candidates = []
    for position, pdf in enumerate(pdfs, start=1):
        try:
            display_path = pdf.relative_to(fs_folder)
        except ValueError:
            display_path = pdf.name
        print(f"  Inspecting: {position}/{len(pdfs)} {display_path}")
        candidate = make_candidate(pdf, allow_ocr=allow_ocr)
        if candidate is not None:
            candidates.append(candidate)
    if not candidates:
        return None

    latest_year_end = max(candidate.financial_year_end for candidate in candidates)
    latest = [
        candidate
        for candidate in candidates
        if candidate.financial_year_end == latest_year_end
    ]
    return min(latest, key=candidate_stage)


def read_selected_pdf(
    pdf_path: Path,
    allow_ocr: bool = False,
) -> tuple[PdfReader, list[str]]:
    reader = PdfReader(str(pdf_path))
    pages = [normalize_text(page.extract_text() or "") for page in reader.pages]
    weak_pages = [index for index, text in enumerate(pages) if len(flat_text(text)) < 20]
    if allow_ocr and weak_pages:
        recovered = ocr_pages(pdf_path, weak_pages)
        for index, text in recovered.items():
            if len(flat_text(text)) > len(flat_text(pages[index])):
                pages[index] = text
    return reader, pages


def extract_date_fields(text: str) -> dict[str, date | None]:
    labels = {
        COLUMNS[2]: r"Company's\s+current\s+financial\s+year\s+start\s+date",
        COLUMNS[3]: r"Company's\s+current\s+financial\s+year\s+end\s+date",
        COLUMNS[4]: r"Date\s+of\s+financial\s+statements\s+approved\s+by\s+Board\s+of\s+Directors",
        COLUMNS[5]: r"Date\s+of\s+circulation\s+of\s+financial\s+statements\s+and\s+reports\s+to\s+members",
        COLUMNS[6]: r"Date\s+of\s+Statutory\s+Declaration",
    }
    return {column: date_after_label(text, pattern) for column, pattern in labels.items()}


def clean_name(value: str | None) -> str | None:
    if not value:
        return None
    value = re.sub(r"\s+", " ", value).strip(" ,;:-")
    value = re.sub(
        r"\s+Company\s+No\s*:\s*\S+.*$",
        "",
        value,
        flags=re.IGNORECASE,
    ).strip()
    value = re.sub(
        r"\s+Page\s+\d+\s+of\s+\d+.*$",
        "",
        value,
        flags=re.IGNORECASE,
    ).strip()
    value = re.sub(
        r"\s+Detailed\s+address\s+of\s+audit\s+firm.*$",
        "",
        value,
        flags=re.IGNORECASE,
    ).strip()
    if not value or value.casefold() in {"nil", "none", "n/a"}:
        return None
    if "tooltiptext" in value.casefold() or value.startswith("#T"):
        return None
    return value


def labelled_director_name(text: str, ordinal: str) -> str | None:
    searchable = flat_text(text)
    match = re.search(
        rf"Name\s+of\s+{ordinal}\s+director\s+who\s+signed\s+(?:the\s+)?Statement\s+by\s+Directors\s*"
        rf"(.*?)"
        rf"(?=Disclosure\s+whether|Type\s+of\s+identification|Identification\s+number|"
        rf"Name\s+of\s+(?:first|second|third|fourth|fifth)\s+director|Date\s+of\s+signing|$)",
        searchable,
        re.IGNORECASE,
    )
    return clean_name(match.group(1)) if match else None


def conventional_statement_signers(text: str) -> tuple[str | None, str | None]:
    searchable = flat_text(text)
    match = re.search(
        r"STATEMENT\s+BY\s+DIRECTORS.*?\bWe,\s*(.*?)\s+and\s+(.*?),\s+being\s+(?:all\s+)?the\s+directors",
        searchable,
        re.IGNORECASE,
    )
    if not match:
        return None, None
    return clean_name(match.group(1)), clean_name(match.group(2))


def extract_statement_signers(text: str) -> tuple[int | None, str | None, str | None]:
    searchable = flat_text(text)
    count_match = re.search(
        r"Number\s+of\s+directors\s+signing\s+Statement\s+by\s+Directors\s+(\d+)",
        searchable,
        re.IGNORECASE,
    )
    count = int(count_match.group(1)) if count_match else None
    first = labelled_director_name(text, "first")
    second = labelled_director_name(text, "second")

    fallback_first, fallback_second = conventional_statement_signers(text)
    first = first or fallback_first
    second = second or fallback_second
    if count is None and first:
        count = 2 if second else 1
    if count is not None and count < 2:
        second = None
    return count, first, second


def statutory_section(text: str) -> str | None:
    searchable = flat_text(text)
    matches = list(re.finditer(r"STATUTORY\s+DECLARATION", searchable, re.IGNORECASE))
    for match in matches:
        section = searchable[match.start() : match.start() + 5000]
        if re.search(r"Section\s+251\s*\(\s*1\s*\)\s*\(\s*b\s*\)", section, re.IGNORECASE):
            end = re.search(r"\bStatement\s+by\s+Directors\b", section[100:], re.IGNORECASE)
            return section[: 100 + end.start()] if end else section
    return None


def extract_declarant_name(
    text: str,
    first_director: str | None,
    second_director: str | None,
) -> str | None:
    section = statutory_section(text)
    if not section:
        return None

    direct = re.search(
        r"\bI,\s*([A-Z][A-Z .@'&/-]{2,}?)\s*(?:\(|,\s*being|being\s+(?:a|the)\s+director)",
        section,
        re.IGNORECASE,
    )
    candidate = clean_name(direct.group(1)) if direct else None
    if candidate:
        for known_name in (first_director, second_director):
            if known_name and flat_text(candidate).casefold() == flat_text(known_name).casefold():
                return known_name
        return candidate

    subscribed = re.search(
        r"the\s+abovenamed\s+([A-Z][A-Za-z .@'&/-]{2,}?)\s+at\b",
        section,
        re.IGNORECASE,
    )
    candidate = clean_name(subscribed.group(1)) if subscribed else None
    if candidate:
        for known_name in (first_director, second_director):
            if known_name and flat_text(candidate).casefold() == flat_text(known_name).casefold():
                return known_name
        return candidate

    whole = flat_text(text)
    first_is_responsible = re.search(
        r"Disclosure\s+whether\s+the\s+first\s+director.*?Primarily\s+responsible\s+for\s+financial\s+management",
        whole,
        re.IGNORECASE,
    )
    if first_is_responsible and first_director:
        return first_director
    second_is_responsible = re.search(
        r"Disclosure\s+whether\s+the\s+second\s+director.*?Primarily\s+responsible\s+for\s+financial\s+management",
        whole,
        re.IGNORECASE,
    )
    if second_is_responsible and second_director:
        return second_director
    return None


def extract_audit_firm(reader: PdfReader, pages: list[str]) -> str | None:
    for page_number, text in enumerate(pages):
        searchable = flat_text(text)
        match = re.search(
            r"Name\s+of\s+audit\s+firm\s+(.*?)\s+Address\s+line\s+1",
            searchable,
            re.IGNORECASE,
        )
        candidate = clean_name(match.group(1)) if match else None
        if candidate and candidate.casefold() not in {"firm", "address"}:
            return candidate

        if not re.search(r"Name\s+of\s+audit\s+firm", searchable, re.IGNORECASE):
            continue
        try:
            layout = normalize_text(
                reader.pages[page_number].extract_text(extraction_mode="layout") or ""
            )
        except Exception:
            continue
        lines = layout.splitlines()
        for index, line in enumerate(lines):
            lower = line.casefold()
            if "name of audit" not in lower:
                continue
            start = lower.index("name of audit")
            header_end = index
            end = None
            for offset in range(0, 3):
                if index + offset >= len(lines):
                    break
                header_line = lines[index + offset].casefold()
                if "address line 1" in header_line:
                    end = header_line.index("address line 1")
                    header_end = index + offset
                    break
            if end is None or end <= start:
                continue
            pieces: list[str] = []
            started = False
            blank_run = 0
            for row in lines[header_end + 1 : header_end + 14]:
                piece = row[start:end].strip() if len(row) > start else ""
                if piece.casefold() in {"firm", "audit firm"}:
                    continue
                if piece:
                    started = True
                    blank_run = 0
                    pieces.append(piece)
                elif started:
                    blank_run += 1
                    if blank_run >= 2:
                        break
            candidate = clean_name(" ".join(pieces))
            if candidate:
                return candidate

    whole = "\n".join(pages)
    signature = re.search(
        r"(?m)^\s*([A-Z][A-Z .&'-]{3,}?)\s*(?:\(AF\s*[:.]?\s*\d+\)|AF\s*[:.]?\s*\d+)\s*$",
        whole,
    )
    candidate = clean_name(signature.group(1)) if signature else None
    if candidate:
        candidate = re.sub(r"^AUDITORS?\s+", "", candidate, flags=re.IGNORECASE)
    return clean_name(candidate)


def parse_money(value: str) -> float | int | None:
    value = value.strip()
    if not value:
        return None
    if value in {"-", "--"}:
        return 0
    negative = value.startswith("(") and value.endswith(")")
    cleaned = value.strip("()").replace(",", "")
    try:
        number = float(cleaned)
    except ValueError:
        return None
    if negative:
        number = -number
    return int(number) if number.is_integer() else number


def extract_current_director_fee(reader: PdfReader, pages: list[str]) -> float | int | None:
    for page_number, text in enumerate(pages):
        searchable = flat_text(text).casefold()
        if "director's remuneration" not in searchable or "fees" not in searchable:
            continue
        try:
            layout = normalize_text(
                reader.pages[page_number].extract_text(extraction_mode="layout") or ""
            )
        except Exception:
            layout = text

        lines = layout.splitlines()
        in_section = False
        consolidated = bool(
            re.search(r"\bGroup\b", layout, re.IGNORECASE)
            and re.search(r"\bCompany\b", layout, re.IGNORECASE)
        )
        for line in lines:
            if re.search(r"Director's\s+remuneration", line, re.IGNORECASE):
                in_section = True
                continue
            if not in_section:
                continue
            if re.search(r"Total\s+Director's\s+remuneration", line, re.IGNORECASE):
                break
            fee_match = re.match(r"^\s*Fees\b(.*)$", line, re.IGNORECASE)
            if not fee_match:
                continue
            remainder = fee_match.group(1).strip()
            if not remainder:
                return None
            tokens = re.findall(r"\(?-?\d[\d,]*(?:\.\d+)?\)?|(?<!\S)-(?!\S)", remainder)
            values = [parse_money(token) for token in tokens]
            values = [value for value in values if value is not None]
            if not values:
                return None
            if consolidated and len(values) >= 3:
                return values[2]
            return values[0]
    return None


def extract_selected(
    candidate: Candidate,
    company: str,
    allow_ocr: bool = False,
) -> dict:
    reader, pages = read_selected_pdf(candidate.path, allow_ocr=allow_ocr)
    whole = "\n".join(pages)
    row = {column: None for column in COLUMNS}
    row["Company"] = company
    row["Source PDF"] = str(candidate.path.resolve())
    row.update(extract_date_fields(whole))

    if row[COLUMNS[3]] is None:
        row[COLUMNS[3]] = candidate.financial_year_end

    count, first, second = extract_statement_signers(whole)
    row[COUNT_COLUMN] = count
    row[COLUMNS[9]] = first
    row[COLUMNS[10]] = second if count and count > 1 else None
    row[COLUMNS[7]] = extract_declarant_name(whole, first, second)
    row[COLUMNS[11]] = extract_audit_firm(reader, pages)
    row[FEE_COLUMN] = extract_current_director_fee(reader, pages)
    return row


def validate_rows(rows: list[dict]) -> None:
    companies = [row["Company"] for row in rows]
    duplicates = sorted({name for name in companies if companies.count(name) > 1})
    if duplicates:
        raise ValueError(f"Duplicate companies in output: {', '.join(duplicates)}")
    for row in rows:
        if list(row) != COLUMNS:
            raise ValueError(f"Unexpected row schema for {row.get('Company', '(unknown)')}")
        count = row[COUNT_COLUMN]
        if count is not None and (not isinstance(count, int) or count < 0):
            raise ValueError(f"Invalid director count for {row['Company']}: {count!r}")
        if (count or 0) < 2 and row[COLUMNS[10]]:
            raise ValueError(f"Second director present when count is below two: {row['Company']}")


def write_workbook(rows: list[dict], output_file: Path) -> None:
    validate_rows(rows)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    temp_file = output_file.with_name(f".{output_file.stem}.tmp.xlsx")
    if temp_file.exists():
        temp_file.unlink()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(COLUMNS)
    for row in rows:
        sheet.append([row[column] for column in COLUMNS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 45
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:M{max(1, sheet.max_row)}"
    sheet.sheet_view.showGridLines = False

    widths = [34, 65, 19, 19, 22, 22, 19, 34, 18, 34, 34, 34, 22]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    for column in range(3, 8):
        for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=2):
            for item in cell:
                item.number_format = "yyyy-mm-dd"
    for cell in sheet["I"][1:]:
        cell.number_format = "0"
    for cell in sheet["M"][1:]:
        cell.number_format = "#,##0.00;[Red]-#,##0.00"

    if rows:
        table = Table(displayName="FSData", ref=f"A1:M{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    try:
        workbook.save(temp_file)
        check = load_workbook(temp_file, read_only=True, data_only=False)
        try:
            if check.sheetnames != [SHEET_NAME]:
                raise RuntimeError(f"Unexpected worksheets: {check.sheetnames}")
            check_sheet = check[SHEET_NAME]
            headers = [cell.value for cell in next(check_sheet.iter_rows(min_row=1, max_row=1))]
            if headers != COLUMNS:
                raise RuntimeError("Workbook header verification failed")
            if check_sheet.max_row != len(rows) + 1:
                raise RuntimeError("Workbook row-count verification failed")
        finally:
            check.close()
        os.replace(temp_file, output_file)
    finally:
        if temp_file.exists():
            temp_file.unlink()


def process_clients(
    client_root: Path,
    allow_ocr: bool = False,
) -> tuple[list[dict], list[str]]:
    rows: list[dict] = []
    warnings_list: list[str] = []
    if not client_root.is_dir():
        raise FileNotFoundError(f"Client root not found: {client_root}")

    companies = sorted(
        (path for path in client_root.iterdir() if path.is_dir()),
        key=lambda path: path.name.casefold(),
    )
    for company in companies:
        fs_folder = find_fs_folder(company)
        if fs_folder is None:
            continue
        print(f"Processing: {company.name}")
        candidate = select_latest_pdf(fs_folder, allow_ocr=allow_ocr)
        if candidate is None:
            warning = f"{company.name}: no financial-statement PDF with a usable year end"
            warnings_list.append(warning)
            print(f"  WARNING: {warning}")
            continue
        try:
            print(
                f"  Extracting: {candidate.path.name} "
                f"({candidate.page_count} pages)"
            )
            row = extract_selected(
                candidate,
                company.name,
                allow_ocr=allow_ocr,
            )
        except Exception as error:
            warning = f"{company.name}: selected PDF failed ({candidate.path.name}): {error}"
            warnings_list.append(warning)
            print(f"  WARNING: {warning}")
            row = {column: None for column in COLUMNS}
            row["Company"] = company.name
            row["Source PDF"] = str(candidate.path.resolve())
            row[COLUMNS[3]] = candidate.financial_year_end
        rows.append(row)
        missing = [column for column in COLUMNS[2:] if row[column] is None]
        print(f"  Selected: {candidate.path.name} ({candidate.financial_year_end})")
        if missing:
            warning = f"{company.name}: {len(missing)} requested value(s) unavailable"
            warnings_list.append(warning)
            print(f"  WARNING: {warning}")
    return rows, warnings_list


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--client-root", type=Path, default=CLIENT_ROOT)
    parser.add_argument("--output", type=Path, default=OUTPUT_FILE)
    parser.add_argument(
        "--ocr",
        action="store_true",
        help="OCR sparse pages in scanned PDFs (disabled by default)",
    )
    return parser.parse_args()


def main() -> None:
    arguments = parse_arguments()
    print("Financial Statements Extractor")
    print(f"Input : {arguments.client_root}")
    print(f"Output: {arguments.output}")
    print(f"OCR   : {'enabled' if arguments.ocr else 'disabled'}")
    if arguments.ocr:
        print("OCR   : initializing opt-in engine...")
        get_ocr_engine()
        print("OCR   : engine ready")
    rows, warnings_list = process_clients(
        arguments.client_root,
        allow_ocr=arguments.ocr,
    )
    if not rows:
        raise RuntimeError("No financial-statement rows were produced")
    write_workbook(rows, arguments.output)
    print(f"Rows    : {len(rows)}")
    print(f"Warnings: {len(warnings_list)}")
    print("Workbook: verified")


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        raise SystemExit(f"ERROR: {error}") from error
