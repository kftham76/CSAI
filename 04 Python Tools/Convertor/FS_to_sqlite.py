"""Convert the Financial Statements workbook into a verified SQLite database."""

from __future__ import annotations

import argparse
from datetime import date, datetime
import math
import os
from pathlib import Path
import sqlite3

from openpyxl import load_workbook


SOURCE_FILE = Path(r"D:\CSAI_DATA\Database\FS.Xlsx")
SHEET_NAME = "FS"
DATABASE_FILE = Path(r"C:\CSAI_OS\06 Data\databases\FS.db")
TABLE_NAME = "FS"

COLUMNS = [
    "Company",
    "Source PDF",
    "Company's current financial year start date",
    "Company's current financial year end date",
    "Date of financial statements approved by Board of Directors",
    "Date of circulation of financial statements and reports to members",
    "Date of Statutory Declaration",
    "Statutory Declaration - Name of director who made declaration",
    "Number of directors signing Statement by Directors",
    "Name of first director who signed Statement by Directors",
    "Name of second director who signed Statement by Directors",
    "Name of audit firm",
    "Director's remuneration - Fees (Current Financial Year)",
]

DATE_COLUMNS = set(COLUMNS[2:7])
COUNT_COLUMN = "Number of directors signing Statement by Directors"
FEE_COLUMN = "Director's remuneration - Fees (Current Financial Year)"


def quote(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def normalize_blank(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def normalize_date(value, column: str) -> str | None:
    value = normalize_blank(value)
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        for pattern in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(value, pattern).date().isoformat()
            except ValueError:
                continue
    raise ValueError(f"Invalid date in {column}: {value!r}")


def normalize_count(value) -> int | None:
    value = normalize_blank(value)
    if value is None:
        return None
    if isinstance(value, bool):
        raise ValueError(f"Invalid director count: {value!r}")
    try:
        number = int(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"Invalid director count: {value!r}") from error
    if number < 0 or float(value) != number:
        raise ValueError(f"Invalid director count: {value!r}")
    return number


def normalize_fee(value) -> float | int | None:
    value = normalize_blank(value)
    if value is None:
        return None
    if isinstance(value, bool):
        raise ValueError(f"Invalid director fee: {value!r}")
    if isinstance(value, str):
        if value == "-":
            return 0
        value = value.replace(",", "")
        if value.startswith("(") and value.endswith(")"):
            value = "-" + value[1:-1]
    try:
        number = float(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"Invalid director fee: {value!r}") from error
    if not math.isfinite(number):
        raise ValueError(f"Invalid director fee: {value!r}")
    return int(number) if number.is_integer() else number


def load_rows(source_file: Path) -> list[tuple]:
    if not source_file.is_file():
        raise FileNotFoundError(f"Source workbook not found: {source_file}")
    workbook = load_workbook(source_file, read_only=True, data_only=True)
    try:
        if workbook.sheetnames != [SHEET_NAME]:
            raise ValueError(
                f"Expected only worksheet {SHEET_NAME!r}; found {workbook.sheetnames}"
            )
        sheet = workbook[SHEET_NAME]
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = list(next(iterator))
        except StopIteration as error:
            raise ValueError("The FS worksheet is empty") from error
        if headers != COLUMNS:
            raise ValueError(
                "Unexpected FS worksheet columns. "
                f"Expected {COLUMNS}; found {headers}"
            )

        rows: list[tuple] = []
        companies: set[str] = set()
        for row_number, raw in enumerate(iterator, 2):
            if all(normalize_blank(value) is None for value in raw):
                continue
            values = dict(zip(COLUMNS, raw))
            company = normalize_blank(values["Company"])
            if not isinstance(company, str):
                raise ValueError(f"Row {row_number} has no valid Company")
            if company in companies:
                raise ValueError(f"Duplicate Company at row {row_number}: {company}")
            companies.add(company)

            normalized = []
            for column in COLUMNS:
                value = values[column]
                if column in DATE_COLUMNS:
                    value = normalize_date(value, column)
                elif column == COUNT_COLUMN:
                    value = normalize_count(value)
                elif column == FEE_COLUMN:
                    value = normalize_fee(value)
                else:
                    value = normalize_blank(value)
                    if value is not None and not isinstance(value, str):
                        value = str(value)
                normalized.append(value)

            count = normalized[COLUMNS.index(COUNT_COLUMN)]
            second = normalized[COLUMNS.index("Name of second director who signed Statement by Directors")]
            if (count or 0) < 2 and second is not None:
                raise ValueError(
                    f"Row {row_number} has a second signer but director count is below two"
                )
            rows.append(tuple(normalized))
        return rows
    finally:
        workbook.close()


def create_database(database_file: Path, rows: list[tuple]) -> None:
    database_file.parent.mkdir(parents=True, exist_ok=True)
    temp_file = database_file.with_name(f".{database_file.stem}.tmp.db")
    if temp_file.exists():
        temp_file.unlink()

    definitions = []
    for column in COLUMNS:
        if column == "Company":
            sql_type = "TEXT PRIMARY KEY"
        elif column == COUNT_COLUMN:
            sql_type = "INTEGER"
        elif column == FEE_COLUMN:
            sql_type = "REAL"
        else:
            sql_type = "TEXT"
        definitions.append(f"{quote(column)} {sql_type}")

    try:
        connection = sqlite3.connect(temp_file)
        try:
            connection.execute(
                f"CREATE TABLE {quote(TABLE_NAME)} ({', '.join(definitions)})"
            )
            placeholders = ", ".join("?" for _ in COLUMNS)
            connection.executemany(
                f"INSERT INTO {quote(TABLE_NAME)} VALUES ({placeholders})",
                rows,
            )
            connection.commit()

            integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
            if integrity != "ok":
                raise RuntimeError(f"SQLite integrity check failed: {integrity}")
            info = connection.execute(
                f"PRAGMA table_info({quote(TABLE_NAME)})"
            ).fetchall()
            imported_columns = [item[1] for item in info]
            if imported_columns != COLUMNS:
                raise RuntimeError("SQLite schema verification failed")
            imported = connection.execute(
                f"SELECT * FROM {quote(TABLE_NAME)} ORDER BY rowid"
            ).fetchall()
            if imported != rows:
                raise RuntimeError("SQLite value verification failed")
        finally:
            connection.close()
        os.replace(temp_file, database_file)
    finally:
        if temp_file.exists():
            temp_file.unlink()


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=SOURCE_FILE)
    parser.add_argument("--database", type=Path, default=DATABASE_FILE)
    return parser.parse_args()


def main() -> None:
    arguments = parse_arguments()
    rows = load_rows(arguments.source)
    create_database(arguments.database, rows)
    print(f"Source   : {arguments.source}")
    print(f"Database : {arguments.database}")
    print(f"Table    : {TABLE_NAME}")
    print(f"Rows     : {len(rows)}")
    print("Integrity: ok")


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        raise SystemExit(f"ERROR: {error}") from error
