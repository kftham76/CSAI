"""Convert the vertical New Incorporation workbook into SQLite."""

from __future__ import annotations

import argparse
from datetime import date, datetime
import os
from pathlib import Path
import shutil
import sqlite3
from typing import Any

from openpyxl import load_workbook


SOURCE_FILE = Path(r"D:\CSAI_DATA\Database\New Incorp.xlsx")
DATABASE_FILE = Path(r"C:\CSAI_OS\06 Data\databases\new_incorp.db")
TABLE_NAME = "New_Incorp"
EXPECTED_COMPANIES = 80
CORE_REQUIRED_FIELDS = (
    "Folder",
    "Company Name",
    "Reg No",
)
# The current vertical workbook schema only guarantees the company identity
# fields below. Extraction-status and UpdatedAt rows are optional content and
# must not prevent conversion when they are absent from the latest workbook.
REQUIRED_FIELDS = CORE_REQUIRED_FIELDS


def clean(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(microsecond=0).isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def read_index(workbook) -> list[dict[str, str]]:
    if "Index" not in workbook.sheetnames:
        raise ValueError("Workbook is missing the Index worksheet")
    sheet = workbook["Index"]
    headers = [clean(cell.value) for cell in sheet[1]]
    required = ("Sequence", "Worksheet", "Folder", "Company Name", "Reg No")
    missing = [field for field in required if field not in headers]
    if missing:
        raise ValueError(f"Index is missing columns: {', '.join(missing)}")
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and str(value).strip() for value in values):
            continue
        item = {str(headers[index]): clean(value) or "" for index, value in enumerate(values) if index < len(headers) and headers[index]}
        rows.append(item)
    if len(rows) != EXPECTED_COMPANIES:
        raise ValueError(f"Expected {EXPECTED_COMPANIES} Index companies, found {len(rows)}")
    worksheet_names = [item["Worksheet"] for item in rows]
    if any(not name for name in worksheet_names) or len(set(worksheet_names)) != len(worksheet_names):
        raise ValueError("Index Worksheet values must be nonblank and unique")
    return rows


def read_company_sheet(
    sheet,
    *,
    min_row: int = 3,
    required_fields: tuple[str, ...] = REQUIRED_FIELDS,
) -> tuple[list[str], dict[str, str | None]]:
    merged_start_rows = {
        merged.min_row
        for merged in sheet.merged_cells.ranges
        if merged.min_col == 1 and merged.max_col >= 3
    }
    labels: list[str] = []
    row: dict[str, str | None] = {}
    for row_number in range(min_row, sheet.max_row + 1):
        if row_number in merged_start_rows:
            continue
        label = clean(sheet.cell(row_number, 1).value)
        if not label:
            continue
        if label in row:
            raise ValueError(f"{sheet.title}: duplicate field label {label!r}")
        value = clean(sheet.cell(row_number, 2).value)
        auxiliary = clean(sheet.cell(row_number, 3).value)
        if auxiliary:
            value = f"{value or ''} | {auxiliary}".strip(" |")
        labels.append(label)
        row[label] = value
    missing = [field for field in required_fields if field not in row]
    if missing:
        raise ValueError(f"{sheet.title}: missing required fields: {', '.join(missing)}")
    return labels, row


def load_workbook_rows(
    source_file: Path,
) -> tuple[list[str], list[dict[str, str | None]], bool]:
    if not source_file.is_file():
        raise FileNotFoundError(f"Source workbook not found: {source_file}")
    workbook = load_workbook(source_file, data_only=True, read_only=False)
    try:
        if "Index" in workbook.sheetnames:
            index = read_index(workbook)
            expected_labels: list[str] | None = None
            rows = []
            for item in index:
                sheet_name = item["Worksheet"]
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"Index references missing worksheet: {sheet_name}")
                labels, row = read_company_sheet(workbook[sheet_name])
                if expected_labels is None:
                    expected_labels = labels
                elif labels != expected_labels:
                    raise ValueError(
                        f"{sheet_name}: vertical field schema differs from the first company sheet"
                    )
                if row["Folder"] != item["Folder"]:
                    raise ValueError(f"{sheet_name}: Folder does not agree with Index")
                if row["Reg No"] != item["Reg No"]:
                    raise ValueError(f"{sheet_name}: Reg No does not agree with Index")
                rows.append(row)

            columns = expected_labels or []
            if len(rows) != EXPECTED_COMPANIES:
                raise ValueError(f"Expected {EXPECTED_COMPANIES} company rows, found {len(rows)}")
            for key in ("Folder", "Reg No"):
                values = [row.get(key) for row in rows]
                if any(not value for value in values) or len(set(values)) != len(values):
                    raise ValueError(f"{key} values must be nonblank and unique")
            return columns, rows, True

        if len(workbook.worksheets) != 1:
            available = ", ".join(workbook.sheetnames) or "(none)"
            raise ValueError(
                "A workbook without an Index worksheet must contain exactly one "
                f"vertical company worksheet; found: {available}"
            )
        columns, row = read_company_sheet(
            workbook.worksheets[0],
            min_row=1,
            required_fields=CORE_REQUIRED_FIELDS,
        )
        return columns, [row], False
    finally:
        workbook.close()


def quote_identifier(value: str) -> str:
    return "[" + value.replace("]", "]]" ) + "]"


def create_table(connection: sqlite3.Connection, columns: list[str], rows: list[dict[str, str | None]], table_name: str) -> None:
    quoted_table = quote_identifier(table_name)
    connection.execute(f"DROP TABLE IF EXISTS {quoted_table}")
    definitions = ", ".join(f"{quote_identifier(column)} TEXT" for column in columns)
    connection.execute(f"CREATE TABLE {quoted_table} ({definitions})")
    placeholders = ", ".join("?" for _ in columns)
    names = ", ".join(quote_identifier(column) for column in columns)
    connection.executemany(
        f"INSERT INTO {quoted_table} ({names}) VALUES ({placeholders})",
        [[row.get(column) for column in columns] for row in rows],
    )


def verify_table(connection: sqlite3.Connection, columns: list[str], rows: list[dict[str, str | None]], table_name: str) -> None:
    quoted_table = quote_identifier(table_name)
    info = list(connection.execute(f"PRAGMA table_info({quoted_table})"))
    actual_columns = [item[1] for item in info]
    actual_types = [item[2].upper() for item in info]
    if actual_columns != columns:
        raise RuntimeError("SQLite column validation failed")
    if actual_types != ["TEXT"] * len(columns):
        raise RuntimeError("Every New_Incorp column must have TEXT affinity")
    actual_rows = list(connection.execute(f"SELECT * FROM {quoted_table} ORDER BY rowid"))
    expected_rows = [tuple(row.get(column) for column in columns) for row in rows]
    if actual_rows != expected_rows:
        raise RuntimeError("SQLite value validation failed")
    integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
    if integrity != "ok":
        raise RuntimeError(f"SQLite integrity check failed: {integrity}")


def table_exists(connection: sqlite3.Connection, table_name: str) -> bool:
    return connection.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table_name,),
    ).fetchone() is not None


def matching_rowids(
    connection: sqlite3.Connection,
    row: dict[str, str | None],
    table_name: str,
) -> set[int]:
    quoted_table = quote_identifier(table_name)
    matches: set[int] = set()
    for key in ("Reg No", "Folder"):
        value = row.get(key)
        if not value:
            continue
        found = connection.execute(
            f"SELECT rowid FROM {quoted_table} WHERE {quote_identifier(key)} = ?",
            (value,),
        ).fetchall()
        if len(found) > 1:
            raise ValueError(f"Database contains duplicate {key} value: {value}")
        if found:
            matches.add(found[0][0])
    if len(matches) > 1:
        raise ValueError("Folder and Reg No match different database rows")
    return matches


def upsert_company(
    connection: sqlite3.Connection,
    columns: list[str],
    row: dict[str, str | None],
    table_name: str,
) -> str:
    quoted_table = quote_identifier(table_name)
    if not table_exists(connection, table_name):
        create_table(connection, columns, [row], table_name)
        return "inserted"

    info = list(connection.execute(f"PRAGMA table_info({quoted_table})"))
    existing_columns = [item[1] for item in info]
    for column in columns:
        if column not in existing_columns:
            connection.execute(
                f"ALTER TABLE {quoted_table} ADD COLUMN {quote_identifier(column)} TEXT"
            )
            existing_columns.append(column)

    matches = matching_rowids(connection, row, table_name)
    if matches:
        rowid = next(iter(matches))
        # A single-company workbook is the complete current snapshot for that
        # company. Clear database fields that are absent from the latest Excel
        # sheet instead of preserving stale values from an older conversion.
        assignments = ", ".join(
            f"{quote_identifier(column)} = ?" for column in existing_columns
        )
        connection.execute(
            f"UPDATE {quoted_table} SET {assignments} WHERE rowid = ?",
            [row.get(column) for column in existing_columns] + [rowid],
        )
        return "updated"

    names = ", ".join(quote_identifier(column) for column in columns)
    placeholders = ", ".join("?" for _ in columns)
    connection.execute(
        f"INSERT INTO {quoted_table} ({names}) VALUES ({placeholders})",
        [row.get(column) for column in columns],
    )
    return "inserted"


def verify_upsert(
    connection: sqlite3.Connection,
    columns: list[str],
    row: dict[str, str | None],
    table_name: str,
) -> None:
    matches = matching_rowids(connection, row, table_name)
    if len(matches) != 1:
        raise RuntimeError("SQLite company upsert validation failed")
    rowid = next(iter(matches))
    info = list(
        connection.execute(
            f"PRAGMA table_info({quote_identifier(table_name)})"
        )
    )
    database_columns = [item[1] for item in info]
    names = ", ".join(quote_identifier(column) for column in database_columns)
    actual = connection.execute(
        f"SELECT {names} FROM {quote_identifier(table_name)} WHERE rowid = ?",
        (rowid,),
    ).fetchone()
    expected = tuple(row.get(column) for column in database_columns)
    if actual != expected:
        raise RuntimeError("SQLite company value validation failed")
    integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
    if integrity != "ok":
        raise RuntimeError(f"SQLite integrity check failed: {integrity}")


def replace_live_table(columns: list[str], rows: list[dict[str, str | None]], database_file: Path) -> None:
    staging = f"__sync_{TABLE_NAME}"
    connection = sqlite3.connect(database_file, timeout=30)
    try:
        with connection:
            create_table(connection, columns, rows, staging)
            verify_table(connection, columns, rows, staging)
        try:
            connection.execute("BEGIN IMMEDIATE")
            connection.execute(f"DROP TABLE IF EXISTS {quote_identifier(TABLE_NAME)}")
            connection.execute(f"ALTER TABLE {quote_identifier(staging)} RENAME TO {quote_identifier(TABLE_NAME)}")
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        verify_table(connection, columns, rows, TABLE_NAME)
    finally:
        try:
            connection.execute(f"DROP TABLE IF EXISTS {quote_identifier(staging)}")
            connection.commit()
        finally:
            connection.close()


def write_database(columns: list[str], rows: list[dict[str, str | None]], database_file: Path) -> None:
    database_file.parent.mkdir(parents=True, exist_ok=True)
    temp_file = database_file.with_name(f".{database_file.stem}.{os.getpid()}.tmp.db")
    if temp_file.exists():
        temp_file.unlink()
    if database_file.exists():
        shutil.copy2(database_file, temp_file)
    try:
        connection = sqlite3.connect(temp_file)
        try:
            with connection:
                create_table(connection, columns, rows, TABLE_NAME)
            verify_table(connection, columns, rows, TABLE_NAME)
        finally:
            connection.close()
        try:
            os.replace(temp_file, database_file)
        except PermissionError:
            replace_live_table(columns, rows, database_file)
    finally:
        if temp_file.exists():
            temp_file.unlink()


def upsert_database(
    columns: list[str],
    row: dict[str, str | None],
    database_file: Path,
) -> str:
    database_file.parent.mkdir(parents=True, exist_ok=True)
    temp_file = database_file.with_name(f".{database_file.stem}.{os.getpid()}.tmp.db")
    if temp_file.exists():
        temp_file.unlink()
    if database_file.exists():
        shutil.copy2(database_file, temp_file)
    action = "inserted"
    try:
        connection = sqlite3.connect(temp_file, timeout=30)
        try:
            with connection:
                action = upsert_company(connection, columns, row, TABLE_NAME)
            verify_upsert(connection, columns, row, TABLE_NAME)
        finally:
            connection.close()
        try:
            os.replace(temp_file, database_file)
        except PermissionError:
            connection = sqlite3.connect(database_file, timeout=30)
            try:
                with connection:
                    action = upsert_company(connection, columns, row, TABLE_NAME)
                verify_upsert(connection, columns, row, TABLE_NAME)
            finally:
                connection.close()
    finally:
        if temp_file.exists():
            temp_file.unlink()
    return action


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=SOURCE_FILE)
    parser.add_argument("--database", type=Path, default=DATABASE_FILE)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv)
    columns, rows, replace_all = load_workbook_rows(args.source)
    if replace_all:
        write_database(columns, rows, args.database)
        action = "replaced"
        mode = "indexed workbook"
    else:
        action = upsert_database(columns, rows[0], args.database)
        mode = "single vertical worksheet"
    print(f"Source    : {args.source}")
    print(f"Database  : {args.database}")
    print(f"Table     : {TABLE_NAME}")
    print(f"Mode      : {mode}")
    print(f"Action    : {action}")
    print(f"Rows      : {len(rows)}")
    print(f"Columns   : {len(columns)}")
    print("Integrity : ok")


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        raise SystemExit(f"ERROR: {error}") from error
